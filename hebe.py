import openpyxl as xl
import re

# open 常规运单模板.xlsx and do assignments
shipping_model_file = xl.load_workbook("常规运单模板.xlsx")
shmo_sheet = shipping_model_file["Sheet1"]

# open WMS_Pacrel_Outbound_Order_20221003.xlsx and do assignments
order_file = xl.load_workbook("WMS_Pacrel_Outbound_Order_20221003.xlsx")
order_sheet = order_file["0"]

dimensions_pattern = r"([0-9]*\.[0-9]*)\("

for row in range(2, order_sheet.max_row + 1):
    shmo_sheet.cell(row, 1).value = row - 1
    shmo_sheet.cell(row, 2).value = "USPS1"
    shmo_sheet.cell(row, 3).value = order_sheet.cell(row, 1).value
    shmo_sheet.cell(row, 5).value = "XMSR"
    shmo_sheet.cell(row, 6).value = "CSS"
    shmo_sheet.cell(row, 7).value = "909-259-9277"
    shmo_sheet.cell(row, 8).value = "CA"
    shmo_sheet.cell(row, 9).value = "Ontario"
    shmo_sheet.cell(row, 10).value = "1550 S. Grove Ave."
    shmo_sheet.cell(row, 11).value = "91761"
    shmo_sheet.cell(row, 12).value = order_sheet.cell(row, 29).value
    shmo_sheet.cell(row, 14).value = "000000000"
    shmo_sheet.cell(row, 15).value = order_sheet.cell(row, 24).value
    shmo_sheet.cell(row, 16).value = order_sheet.cell(row, 25).value
    shmo_sheet.cell(row, 17).value = order_sheet.cell(row, 30).value
    shmo_sheet.cell(row, 18).value = order_sheet.cell(row, 31).value
    shmo_sheet.cell(row, 19).value = order_sheet.cell(row, 28).value
    shmo_sheet.cell(row, 20).value = "lb/in"
    shmo_sheet.cell(row, 21).value = order_sheet.cell(row, 40).value

    # 4 lines below: separate and convert cm to inch
    dimensions = re.split(r"\*", order_sheet.cell(row, 37).value)
    inch_len = float(re.search(dimensions_pattern, dimensions[0])[1]) / 2.54 + 0.5
    inch_w = float(re.search(dimensions_pattern, dimensions[1])[1]) / 2.54 + 0.5
    inch_h = float(re.search(dimensions_pattern, dimensions[2])[1]) / 2.54 + 0.5

    shmo_sheet.cell(row, 22).value = round(inch_len, 2)
    shmo_sheet.cell(row, 23).value = round(inch_w, 2)
    shmo_sheet.cell(row, 24).value = round(inch_h, 2)
    shmo_sheet.cell(row, 25).value = 1
    shmo_sheet.cell(row, 26).value = str(order_sheet.cell(row, 11).value) + "*" + str(order_sheet.cell(row, 12).value)

shipping_model_file.save("result.xlsx")
